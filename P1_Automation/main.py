import openpyxl as xl
import os

wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)

wb = xl.load_workbook('transactions.xlsx')

for row in range(1, sheet.max_row + 1):
    for col in range(1, 6):
        cell = sheet.cell(row, col)
        print(cell.value)
    print("\n")
for row in range(2,sheet.max_row+1):
    #print(row)
    cell=sheet.cell(row,3)
    #print(cell.value)
    corrected_price=cell.value*0.9
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=corrected_price

wb.save('transection2.xslx')
if os.path.exists('transection2.xslx'):
    print("Saved!!!!")
else:
    print("Not Saved!!!!!")






